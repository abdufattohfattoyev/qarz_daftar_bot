"""
Hisobot generatorlari — chiroyli Excel va rasm (PNG).
"""
from io import BytesIO
from decimal import Decimal
from django.utils import timezone
from django.db.models import Sum, Q

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import Debt

GREEN = '16A34A'
RED = 'EF4444'
DARK = '0F172A'
GRAY = '64748B'
LIGHT = 'F1F5F9'


def _balances(user):
    """Har valyuta bo'yicha (berdim_qoldi, oldim_qoldi) qaytaradi."""
    out = {}
    for cur in ('UZS', 'USD'):
        a = Debt.objects.filter(
            user=user, currency=cur, status__in=['active', 'partial']
        ).aggregate(
            gs=Sum('amount',      filter=Q(debt_type='gave')),
            gp=Sum('paid_amount', filter=Q(debt_type='gave')),
            rs=Sum('amount',      filter=Q(debt_type='got')),
            rp=Sum('paid_amount', filter=Q(debt_type='got')),
        )
        gave = (a['gs'] or Decimal(0)) - (a['gp'] or Decimal(0))
        got  = (a['rs'] or Decimal(0)) - (a['rp'] or Decimal(0))
        if gave or got:
            out[cur] = (gave, got)
    return out


# ── EXCEL ────────────────────────────────────────────────────────────────────

def build_excel(user):
    debts = Debt.objects.filter(user=user).select_related('contact').order_by('-created_at')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Qarz Yordamchi'

    thin = Side(style='thin', color='E2E8F0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    name = user.full_name or user.telegram_username or 'Foydalanuvchi'

    # ── Sarlavha ──
    ws.merge_cells('A1:K1')
    c = ws['A1']
    c.value = '📒 QARZ YORDAMCHI — HISOBOT'
    c.font = Font(bold=True, size=16, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor=GREEN)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:K2')
    c = ws['A2']
    c.value = f'{name}  ·  {timezone.now().strftime("%d.%m.%Y %H:%M")}'
    c.font = Font(size=10, color=GRAY)
    c.alignment = Alignment(horizontal='center')

    # ── Balans bloki ──
    row = 4
    for cur, (gave, got) in _balances(user).items():
        net = gave - got
        ws.cell(row, 1, f'{cur} — Menga berishadi').font = Font(bold=True, color=GREEN)
        ws.cell(row, 3, float(gave)).number_format = '#,##0'
        ws.cell(row, 4, f'Men beraman').font = Font(bold=True, color=RED)
        ws.cell(row, 6, float(got)).number_format = '#,##0'
        ws.cell(row, 7, 'Sof balans').font = Font(bold=True, color=DARK)
        nc = ws.cell(row, 9, float(net))
        nc.number_format = '#,##0'
        nc.font = Font(bold=True, color=GREEN if net >= 0 else RED)
        row += 1

    header_row = row + 1

    # ── Jadval sarlavhasi ──
    headers = ['Sana', 'Kontakt', 'Telefon', 'Tur', 'Miqdor',
               "To'langan", 'Qoldi', 'Valyuta', 'Holat', 'Izoh', 'Muddat']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(header_row, col, h)
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill('solid', fgColor=DARK)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    ws.row_dimensions[header_row].height = 22

    # ── Ma'lumotlar ──
    type_label = {'gave': 'Men berdim', 'got': 'Men oldim'}
    status_label = {'active': 'Faol', 'partial': 'Qisman', 'paid': "To'langan"}
    status_color = {'active': RED, 'partial': 'F97316', 'paid': GREEN}

    r = header_row + 1
    for debt in debts:
        ws.cell(r, 1, debt.created_at.strftime('%d.%m.%Y %H:%M'))
        ws.cell(r, 2, debt.contact.name)
        ws.cell(r, 3, debt.contact.phone or '—')
        tc = ws.cell(r, 4, type_label.get(debt.debt_type, ''))
        tc.font = Font(color=GREEN if debt.debt_type == 'gave' else RED, bold=True)
        ws.cell(r, 5, float(debt.amount)).number_format = '#,##0'
        ws.cell(r, 6, float(debt.paid_amount)).number_format = '#,##0'
        ws.cell(r, 7, float(debt.remaining_amount)).number_format = '#,##0'
        ws.cell(r, 8, debt.currency)
        sc = ws.cell(r, 9, status_label.get(debt.status, ''))
        sc.font = Font(color=status_color.get(debt.status, GRAY), bold=True)
        ws.cell(r, 10, debt.note or '')
        ws.cell(r, 11, debt.due_date.strftime('%d.%m.%Y') if debt.due_date else '')

        for col in range(1, 12):
            cell = ws.cell(r, col)
            cell.border = border
            if r % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='F8FAFC')
        r += 1

    # ── Brend (pastki qator) ──
    brand_row = r + 1
    ws.merge_cells(start_row=brand_row, start_column=1, end_row=brand_row, end_column=11)
    bc = ws.cell(brand_row, 1, '📒 Qarz Yordamchi — @Qarz_Yordamchi_Bot')
    bc.font = Font(bold=True, color=GREEN, size=11)
    bc.alignment = Alignment(horizontal='center')

    # ── Ustun kengligi ──
    widths = [17, 20, 16, 12, 13, 13, 13, 9, 11, 24, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = ws.cell(header_row + 1, 1)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── RASM (PNG) ───────────────────────────────────────────────────────────────

def build_image(user):
    from PIL import Image, ImageDraw, ImageFont

    def font(size, bold=False):
        path = ('/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf' if bold
                else '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf')
        try:
            return ImageFont.truetype(path, size)
        except Exception:
            return ImageFont.load_default()

    balances = _balances(user)
    name = user.full_name or user.telegram_username or 'Foydalanuvchi'

    W = 800
    H = 280 + len(balances) * 150 + 40   # +40 — pastki brend qatori uchun
    img = Image.new('RGB', (W, H), '#F0F2F5')
    d = ImageDraw.Draw(img)

    # Header
    d.rectangle([0, 0, W, 120], fill='#16a34a')
    d.text((40, 30), 'Qarz Yordamchi', font=font(34, True), fill='#ffffff')
    d.text((40, 78), f'{name} · {timezone.now().strftime("%d.%m.%Y")}',
           font=font(18), fill='#dcfce7')

    y = 160
    if not balances:
        d.text((40, y), "Hozircha faol qarzlar yo'q", font=font(22), fill='#64748b')
    for cur, (gave, got) in balances.items():
        net = gave - got
        # Card
        d.rounded_rectangle([30, y, W - 30, y + 130], radius=18, fill='#ffffff')
        d.text((55, y + 18), cur, font=font(20, True), fill='#0f172a')

        d.text((55, y + 56), 'Menga berishadi', font=font(15), fill='#64748b')
        d.text((55, y + 80), f'+{gave:,.0f}', font=font(26, True), fill='#16a34a')

        d.text((310, y + 56), 'Men beraman', font=font(15), fill='#64748b')
        d.text((310, y + 80), f'-{got:,.0f}', font=font(26, True), fill='#ef4444')

        d.text((560, y + 56), 'Sof balans', font=font(15), fill='#64748b')
        d.text((560, y + 80), f'{"+" if net >= 0 else "-"}{abs(net):,.0f}',
               font=font(26, True), fill='#16a34a' if net >= 0 else '#ef4444')
        y += 150

    # ── Brend (pastda, markazda) ──
    brand = 'Qarz Yordamchi  -  @Qarz_Yordamchi_Bot'
    bf = font(17, True)
    bw = d.textbbox((0, 0), brand, font=bf)[2]
    d.text(((W - bw) // 2, H - 32), brand, font=bf, fill='#16a34a')

    buf = BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return buf.getvalue()
